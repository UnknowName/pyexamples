import os
import sys
import time
from openpyxl import Workbook, load_workbook


def read_excel(filename: str):
    wb = load_workbook(filename=filename)
    sheet = wb.active
    print("{}: 已将EXCEL加载进内存，开始按第一列的值进行拆分".format(time.asctime()))
    result_dir = os.path.join(os.getcwd(), "结果")
    if not os.path.exists(result_dir):
        os.mkdir(result_dir)
    count = dict()
    # 第一行为表头
    headers = next(sheet.values)
    for values in sheet.values:
        if headers == values:
            continue
        if any(values):
            # values type is tuple,根据第一列的值进行分类
            province = values[0]
            count.setdefault(province, []).append(values)
    print("{}: 拆分统计完成，开始分别写入不同文件".format(time.asctime()))
    for region, rows in count.items():
        wb = Workbook()
        sheet = wb.active
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        wb.save(os.path.join(result_dir, "{}.xlsx".format(region)))


if __name__ == '__main__':
    start = time.time()
    if len(sys.argv) == 2:
        excel_file = sys.argv[1]
        if excel_file.endswith(".xlsx"):
            print("{}: 开始将原始文件全部加载进内存,该过程同文件大小正相关".format(time.asctime()))
            read_excel(excel_file)
            print("{}: EXCEL拆分完成，用时{}秒".format(time.asctime(), time.time() - start))
            exit(0)
        else:
            print("目前只支持xlsx格式，其它不支持")
            exit(9)
    else:
        print("使用示例: python excel.py 201908.xlsx")
        exit(9)
